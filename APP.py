import io
import re
import zipfile
import unicodedata
from pathlib import Path
from typing import List, Tuple, Optional

import streamlit as st
from pypdf import PdfReader, PdfWriter
import pdfplumber

# ============ Utilidades de texto / limpieza ============
def sanitize_filename(name: str, max_len: int = 100) -> str:
    name = re.sub(r"[^\w\s\-_.()]", "", name, flags=re.UNICODE).strip()
    name = re.sub(r"\s+", " ", name)
    name = name[:max_len] or "Plan"
    return name

def clean_title_prefixes(name: str) -> str:
    """Quita títulos comunes al inicio (DR., DRA., LIC., ING., MSC., MAG., MTR., PHD, PROF., etc.)."""
    if not name:
        return name
    pattern = r"^\s*(?:dr\.?|dra\.?|lic\.?|ing\.?|msc\.?|m\.?sc\.?|mag\.?|maestr[eo]|master|mtr\.?|ph\.?d\.?|prof\.?)\s+"
    name = re.sub(pattern, "", name, flags=re.IGNORECASE)
    name = re.sub(r"\s+", " ", name).strip()
    return name

def normalize_text(s: str) -> str:
    """Normaliza: sin tildes, minúsculas y espacios colapsados (para comparar)."""
    s = s or ""
    s = s.replace("\u00A0", " ")
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s

# corta un texto ANTES de la palabra GÉNERO/GENERO (insensible a mayúsculas/acentos)
def cut_before_genero(s: str) -> str:
    if not s:
        return s
    parts = re.split(r"\bG[EÉ]NERO\b", s, flags=re.IGNORECASE, maxsplit=1)
    return parts[0].strip() if parts else s.strip()

# ============ Extracción de textos ============
def get_page_texts_for_start(pdf_bytes: bytes, max_lines: int) -> List[str]:
    """Texto por página (solo primeras max_lines) para detectar INICIO."""
    texts = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            txt = page.extract_text() or ""
            lines = txt.splitlines()
            take = lines if max_lines <= 0 else lines[:max_lines]
            texts.append("\n".join(take))
    return texts

def get_text_block_for_name(pdf_bytes: bytes, start_page: int, end_page: int,
                            scan_pages: int, max_lines_per_page: int) -> str:
    """Concatena las primeras max_lines de las primeras scan_pages páginas del rango."""
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        stop = min(end_page, start_page + scan_pages)
        buf = []
        for p in range(start_page, stop):
            txt = pdf.pages[p].extract_text() or ""
            lines = txt.splitlines()
            take = lines if max_lines_per_page <= 0 else lines[:max_lines_per_page]
            buf.extend(take)
    return "\n".join(buf)

# ============ Detección de inicios ============
def detect_starts_by_patterns(page_texts: List[str], patterns: List[str]) -> List[Tuple[int, Optional[str]]]:
    """
    Devuelve lista de (page_index, etiqueta_capturada_o_None).
    Los patrones pueden tener (.+) para capturar un valor.
    Se evalúa línea por línea.
    """
    regexes = [re.compile(pat, re.IGNORECASE) for pat in patterns]
    starts = []
    for i, txt in enumerate(page_texts):
        found_label = None
        for rx in regexes:
            matched = False
            for line in txt.split("\n"):
                m = rx.search(line)
                if m:
                    matched = True
                    if m.lastindex:
                        cand = sanitize_filename(m.group(1))
                        found_label = cand or None
                    else:
                        found_label = None
                    break
            if matched:
                starts.append((i, found_label))
                break
    return starts

def build_ranges_from_starts(total_pages: int, starts: List[Tuple[int, Optional[str]]]) -> List[Tuple[int, int, Optional[str]]]:
    ranges = []
    for k, (p_ini, label_opt) in enumerate(starts):
        p_fin = starts[k + 1][0] if k + 1 < len(starts) else total_pages
        ranges.append((p_ini, p_fin, label_opt))
    return ranges

def build_ranges_every_n(total_pages: int, n: int) -> List[Tuple[int, int, Optional[str]]]:
    ranges = []
    i = 0
    while i < total_pages:
        end = min(i + n, total_pages)
        ranges.append((i, end, None))
        i = end
    return ranges

# ============ Búsqueda específica del NOMBRE ============
def find_prof_name_in_section(pdf_bytes: bytes, start_page: int, end_page: int,
                              scan_pages: int = 2, lines_per_page: int = 60) -> Optional[str]:
    """
    Busca EXACTAMENTE la línea: NOMBRE DEL PROFESOR(A) : <valor>
    Acepta separadores :  -  –  —  =
    Corta el valor antes de la palabra GÉNERO/GENERO si aparece en la misma línea.
    Devuelve <valor> limpio (sin títulos), o None si no lo encuentra.
    """
    label_rx = re.compile(
        r"^\s*NOMBRE\s+DEL\s+PROFESOR\(A\)\s*[:=\-\u2014\u2013]\s*(.+)$",
        re.IGNORECASE
    )

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        stop = min(end_page, start_page + scan_pages)
        for p in range(start_page, stop):
            txt = pdf.pages[p].extract_text() or ""
            for raw in (txt.splitlines()[:lines_per_page]):
                m = label_rx.search(raw)
                if m and m.group(1).strip():
                    name = m.group(1).strip()
                    name = cut_before_genero(name)
                    name = clean_title_prefixes(name)
                    name = sanitize_filename(name)
                    return name if name else None
    return None

# ============ Búsqueda de LAA/DESCARGA ============
def find_laa_descarga_in_section(pdf_bytes: bytes, start_page: int, end_page: int,
                                 scan_pages: int = 2, lines_per_page: int = 60) -> Optional[str]:
    """
    Busca la línea 'LAA/DESCARGA : <numero>' (admite : - – — = como separadores).
    Devuelve el valor (string) que viene después del separador (con signos/espacios).
    Si no lo encuentra, devuelve None.
    """
    rx = re.compile(r"^\s*LAA/DESCARGA\s*[:=\-\u2014\u2013]\s*(.+)$", re.IGNORECASE)

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        stop = min(end_page, start_page + scan_pages)
        for p in range(start_page, stop):
            txt = pdf.pages[p].extract_text() or ""
            for raw in (txt.splitlines()[:lines_per_page]):
                m = rx.search(raw)
                if m and m.group(1).strip():
                    return m.group(1).strip()
    return None

def is_negative_number_string(s: str) -> bool:
    """
    Determina si el string representa un número NEGATIVO
    (guion '-' justo antes de los dígitos).
    Soporta espacios finos: '- 123' también cuenta como negativo.
    """
    if not s:
        return False
    # quitar separadores de miles comunes (espacios, comas) para evaluar
    s_compact = re.sub(r"[ ,\u00A0]", "", s)
    return bool(re.match(r"^-?\d+(\.\d+)?$", s_compact)) and s_compact.startswith("-")

# ============ Excel (openpyxl) ============
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def build_excel_bytes(detalles_rows: List[dict], errores_rows: List[dict]) -> bytes:
    """Crea un Excel con hojas 'detalles', 'resumen' y 'errores' (si hay), usando openpyxl."""
    wb = Workbook()
    # DETALLES
    ws_det = wb.active
    ws_det.title = "detalles"
    headers = ["orden", "archivo", "nombre_detectado", "pagina_inicio_1based", "pagina_fin_1based", "paginas_en_pdf"]
    ws_det.append(headers)
    for row in detalles_rows:
        ws_det.append([row.get(h) for h in headers])
    for col_idx, h in enumerate(headers, 1):
        ws_det.column_dimensions[get_column_letter(col_idx)].width = max(16, len(h) + 2)

    # RESUMEN
    ws_res = wb.create_sheet("resumen")
    ws_res.append(["nombre_detectado", "cantidad_pdfs"])
    counts = {}
    for r in detalles_rows:
        key = r.get("nombre_detectado") or ""
        counts[key] = counts.get(key, 0) + 1
    for name, qty in sorted(counts.items(), key=lambda x: (-x[1], x[0])):
        ws_res.append([name, qty])
    ws_res.column_dimensions["A"].width = 40
    ws_res.column_dimensions["B"].width = 16

    # ERRORES
    if errores_rows:
        ws_err = wb.create_sheet("errores")
        err_headers = ["orden", "nombre_detectado", "valor_laa_descarga", "motivo", "pagina_inicio_1based", "pagina_fin_1based", "paginas_en_pdf"]
        ws_err.append(err_headers)
        for r in errores_rows:
            ws_err.append([r.get(k) for k in err_headers])
        for col_idx, h in enumerate(err_headers, 1):
            ws_err.column_dimensions[get_column_letter(col_idx)].width = max(18, len(h) + 2)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

# ============ Export (ZIP + Excel) ============
def export_zip_and_excel(
    pdf_bytes: bytes,
    ranges: List[Tuple[int, int, Optional[str]]],
    prefix: str,
    scan_pages: int,
    lines_per_page: int,
) -> Tuple[bytes, bytes, List[dict], List[dict]]:
    """
    Devuelve (zip_bytes, excel_bytes, detalles_rows, errores_rows).
    - El ZIP trae solo los PDFs válidos.
    - Excel trae 'detalles', 'resumen' y 'errores' (si hubo).
    - SIN numeración al inicio del nombre del archivo.
    - Si hay colisión de nombres, agrega sufijo _(2), _(3)...
    - Se excluyen del ZIP las secciones con LAA/DESCARGA NEGATIVO.
    """
    reader = PdfReader(io.BytesIO(pdf_bytes))
    detalles_rows = []
    errores_rows = []
    mem_zip = io.BytesIO()
    used_names: dict[str, int] = {}

    with zipfile.ZipFile(mem_zip, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for idx, (start, end, label_opt) in enumerate(ranges, 1):
            # 1) Detectar nombre
            detected = find_prof_name_in_section(
                pdf_bytes, start, end, scan_pages=scan_pages, lines_per_page=lines_per_page
            )
            if not detected and label_opt:
                detected = sanitize_filename(label_opt)
            if not detected:
                detected = f"Plan_{idx:03d}"

            # 2) Leer LAA/DESCARGA
            laa_val = find_laa_descarga_in_section(
                pdf_bytes, start, end, scan_pages=scan_pages, lines_per_page=lines_per_page
            )

            # 3) Validar si negativo
            if laa_val is not None and is_negative_number_string(laa_val):
                # Registrar error y NO agregar al ZIP
                errores_rows.append({
                    "orden": idx,
                    "nombre_detectado": detected,
                    "valor_laa_descarga": laa_val,
                    "motivo": "LAA/DESCARGA negativo",
                    "pagina_inicio_1based": start + 1,
                    "pagina_fin_1based": end,
                    "paginas_en_pdf": end - start,
                })
                continue  # Skip este plan

            # 4) Si pasa validación, escribir PDF
            base = sanitize_filename(f"{prefix}{detected}") if prefix else sanitize_filename(detected)
            final_name = base
            if final_name in used_names:
                used_names[final_name] += 1
                final_name = f"{base}_({used_names[base]})"
            else:
                used_names[final_name] = 1

            writer = PdfWriter()
            for p in range(start, end):
                writer.add_page(reader.pages[p])
            out_bytes = io.BytesIO()
            writer.write(out_bytes)
            out_bytes.seek(0)

            fname = f"{final_name}.pdf"
            zf.writestr(fname, out_bytes.read())

            detalles_rows.append({
                "orden": idx,
                "archivo": fname,
                "nombre_detectado": detected,
                "pagina_inicio_1based": start + 1,
                "pagina_fin_1based": end,
                "paginas_en_pdf": end - start,
            })

    mem_zip.seek(0)
    excel_bytes = build_excel_bytes(detalles_rows, errores_rows)
    return mem_zip.getvalue(), excel_bytes, detalles_rows, errores_rows

# ============ UI ============
st.set_page_config(page_title="PDF Splitter — Profes y Excel", page_icon="📄")
st.title("📄 Dividir PDF, nombrar por 'NOMBRE DEL PROFESOR(A):' y generar Excel")
st.caption("Corta nombres antes de 'GÉNERO'. Filtra 'LAA/DESCARGA' negativo. Descargas simultáneas.")

with st.sidebar:
    st.header("⚙️ Configuración")
    mode = st.radio("Modo de división", ["Por patrones de inicio", "Cada N páginas"])

    # Patrones para detectar inicio de cada “plan”
    default_patterns = "\n".join([
        r"^\s*plan\s+de\s+clase",              # ejemplo
        r"^\s*profesor(?:a)?\s*:\s*(.+)$",     # si justo aquí aparece un nombre
        r"^\s*docente\s*:\s*(.+)$"
    ])
    patterns_text = st.text_area(
        "Patrones de INICIO (uno por línea; opcionalmente con (.+) para capturar).",
        value=default_patterns if mode == "Por patrones de inicio" else "",
        height=110
    )
    header_lines = st.number_input(
        "Líneas a leer por página (INICIO)",
        min_value=0, max_value=120, value=10, step=1
    )

    n_pages = st.number_input(
        "N páginas por bloque (si eliges 'Cada N páginas')",
        min_value=1, max_value=20, value=2, step=1
    )

    st.markdown("---")
    st.subheader("📛 Búsqueda del nombre / LAA")
    st.write("Se busca **exactamente** `NOMBRE DEL PROFESOR(A): ...` y `LAA/DESCARGA: ...` en las primeras páginas de cada sección.")
    scan_pages = st.number_input("Páginas a escanear por sección", 1, 10, 2, 1)
    lines_for_name = st.number_input("Líneas a leer por página (para nombre/LAA)", 5, 120, 60, 5)

    prefix = st.text_input("Prefijo para archivos (opcional)", value="")

file = st.file_uploader("Sube tu PDF", type=["pdf"])

# Mostrar botones de descarga si ya se generaron (sin recalcular)
if "zip_bytes" in st.session_state and "excel_bytes" in st.session_state:
    st.success("Archivos listos para descargar (sin recalcular).")
    st.download_button(
        "⬇️ Descargar ZIP de PDFs",
        data=st.session_state["zip_bytes"],
        file_name=st.session_state.get("zip_name", "planes_split.zip"),
        mime="application/zip",
        key="dl_zip_ready"
    )
    st.download_button(
        "⬇️ Descargar Excel (detalles, resumen, errores)",
        data=st.session_state["excel_bytes"],
        file_name=st.session_state.get("excel_name", "reporte.xlsx"),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_excel_ready"
    )

if file is not None:
    pdf_bytes = file.read()
    reader = PdfReader(io.BytesIO(pdf_bytes))
    total_pages = len(reader.pages)
    st.info(f"PDF cargado: **{file.name}** — {total_pages} páginas")

    if st.button("🔍 Previsualizar cortes"):
        if mode == "Por patrones de inicio":
            if not patterns_text.strip():
                st.error("Agrega al menos un patrón o usa 'Cada N páginas'.")
            else:
                patterns = [p for p in (patterns_text.splitlines()) if p.strip()]
                page_texts = get_page_texts_for_start(pdf_bytes, header_lines)
                starts = detect_starts_by_patterns(page_texts, patterns)
                if not starts:
                    st.warning("No se detectaron INICIOS. Ajusta patrones o usa 'Cada N páginas'.")
                else:
                    ranges = build_ranges_from_starts(total_pages, starts)
                    st.success(f"Detectados {len(ranges)} cortes/secciones.")
                    st.dataframe({
                        "Inicio (pág. 1-based)": [s[0] + 1 for s in starts],
                        "Fin (pág. 1-based)": [r[1] for r in ranges],
                        "Etiqueta capturada": [s[1] or "" for s in starts],
                    })
        else:
            ranges = build_ranges_every_n(total_pages, n_pages)
            st.success(f"Se crearían {len(ranges)} archivos de {n_pages} páginas (último puede variar).")
            st.dataframe({
                "Inicio (pág. 1-based)": [r[0] + 1 for r in ranges],
                "Fin (pág. 1-based)": [r[1] for r in ranges],
                "Etiqueta capturada": [r[2] or "" for r in ranges],
            })

    st.divider()

    if st.button("✂️ Dividir, validar, nombrar y preparar descargas"):
        # Construir los rangos según el modo
        if mode == "Por patrones de inicio":
            patterns = [p for p in (patterns_text.splitlines()) if p.strip()]
            if not patterns:
                st.error("Agrega patrones o usa 'Cada N páginas'.")
                st.stop()
            page_texts = get_page_texts_for_start(pdf_bytes, header_lines)
            starts = detect_starts_by_patterns(page_texts, patterns)
            if not starts:
                st.error("No se detectaron INICIOS. Revisa los patrones.")
                st.stop()
            ranges = build_ranges_from_starts(total_pages, starts)
        else:
            ranges = build_ranges_every_n(total_pages, n_pages)

        # Exportar
        zip_bytes, excel_bytes, detalles_rows, errores_rows = export_zip_and_excel(
            pdf_bytes, ranges, prefix=prefix,
            scan_pages=scan_pages, lines_per_page=lines_for_name
        )

        # Guardar en sesión para descargas simultáneas
        st.session_state["zip_bytes"] = zip_bytes
        st.session_state["excel_bytes"] = excel_bytes
        st.session_state["zip_name"] = f"{Path(file.name).stem}_split.zip"
        st.session_state["excel_name"] = f"{Path(file.name).stem}_reporte.xlsx"

        # Mensajes y tablas
        st.success(f"¡Listo! Generados {len(detalles_rows)} PDFs válidos."
                   + (f" Se excluyeron {len(errores_rows)} por LAA/DESCARGA negativo." if errores_rows else ""))

        # Botones (ambos ya disponibles)
        st.download_button(
            "⬇️ Descargar ZIP de PDFs",
            data=zip_bytes,
            file_name=st.session_state["zip_name"],
            mime="application/zip",
            key="dl_zip_now"
        )
        st.download_button(
            "⬇️ Descargar Excel (detalles, resumen, errores)",
            data=excel_bytes,
            file_name=st.session_state["excel_name"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_excel_now"
        )

        # Mostrar errores (si hubo)
        if errores_rows:
            st.warning("Los errores encontrados (excluidos del ZIP) por LAA/DESCARGA negativo:")
            import pandas as pd  # solo para mostrar tabla (no se usa en requirements)
            st.dataframe(pd.DataFrame(errores_rows))
        else:
            st.info("No se encontraron errores de LAA/DESCARGA negativa.")

st.markdown("---")
with st.expander("❓ Tips"):
    st.markdown(
        """
- Si el nombre aparece más abajo, sube **“Páginas a escanear”** (3–4) o **“Líneas para nombre/LAA”** (80–100).
- Se aceptan separadores `:`, `-`, `–`, `—`, `=` en las etiquetas.
- El ZIP agrega `_(2)`, `_(3)` si hay archivos con nombres repetidos.
"""
    )
